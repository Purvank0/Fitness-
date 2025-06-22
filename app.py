from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from datetime import datetime
import os
import openpyxl
from openpyxl import Workbook

app = Flask(_name_)
app.secret_key = 'your_secret_key_here'

# Excel file configuration
EXCEL_FILE = 'meals.xlsx'

# Ensure Excel file exists with proper headers
def init_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Meals"
        headers = ['id', 'name', 'category', 'ingredients', 'calories', 'protein', 'carbs', 'fat', 'instructions', 'timestamp']
        ws.append(headers)
        wb.save(EXCEL_FILE)

def get_next_id():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    if ws.max_row <= 1:  # Only header row exists
        return 1
    return ws.cell(row=ws.max_row, column=1).value + 1

def get_all_meals():
    meals = []
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]  # Get headers from first row
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        meal = dict(zip(headers, row))
        meals.append(meal)
    return meals

def get_meal_by_id(meal_id):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]  # Get headers from first row
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == meal_id:  # id is first column
            return dict(zip(headers, row))
    return None

def save_meal(meal):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    row = [
        meal['id'],
        meal['name'],
        meal['category'],
        meal['ingredients'],
        meal['calories'],
        meal['protein'],
        meal['carbs'],
        meal['fat'],
        meal['instructions'],
        meal['timestamp']
    ]
    ws.append(row)
    wb.save(EXCEL_FILE)

def update_meal(meal_id, updated_meal):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == meal_id:  # Check id column
            ws.cell(row=row, column=2, value=updated_meal['name'])
            ws.cell(row=row, column=3, value=updated_meal['category'])
            ws.cell(row=row, column=4, value=updated_meal['ingredients'])
            ws.cell(row=row, column=5, value=updated_meal['calories'])
            ws.cell(row=row, column=6, value=updated_meal['protein'])
            ws.cell(row=row, column=7, value=updated_meal['carbs'])
            ws.cell(row=row, column=8, value=updated_meal['fat'])
            ws.cell(row=row, column=9, value=updated_meal['instructions'])
            ws.cell(row=row, column=10, value=updated_meal['timestamp'])
            break
    
    wb.save(EXCEL_FILE)

def delete_meal(meal_id):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Find the row to delete
    row_to_delete = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == meal_id:
            row_to_delete = row
            break
    
    if row_to_delete:
        ws.delete_rows(row_to_delete)
        wb.save(EXCEL_FILE)

def calculate_nutrition_totals(meal_ids):
    meals = get_all_meals()
    totals = {
        'calories': 0,
        'protein': 0,
        'carbs': 0,
        'fat': 0
    }
    
    for meal_id in meal_ids:
        meal = next((m for m in meals if m['id'] == meal_id), None)
        if meal:
            totals['calories'] += int(meal['calories'])
            totals['protein'] += int(meal['protein'])
            totals['carbs'] += int(meal['carbs'])
            totals['fat'] += int(meal['fat'])
    
    return totals

@app.route('/')
def index():
    search_query = request.args.get('search', '')
    category_filter = request.args.get('category', '')
    meals = get_all_meals()
    
    if search_query:
        meals = [meal for meal in meals if search_query.lower() in meal['name'].lower()]
    
    if category_filter:
        meals = [meal for meal in meals if meal['category'].lower() == category_filter.lower()]
    
    categories = set(meal['category'] for meal in get_all_meals())
    
    return render_template('index.html', meals=meals, search_query=search_query, 
                         categories=categories, selected_category=category_filter)

@app.route('/add', methods=['GET', 'POST'])
def add_meal():
    if request.method == 'POST':
        name = request.form['name']
        category = request.form['category']
        ingredients = request.form['ingredients']
        calories = request.form['calories']
        protein = request.form['protein']
        carbs = request.form['carbs']
        fat = request.form['fat']
        instructions = request.form['instructions']
        
        if not name or not category or not ingredients:
            flash('Name, category, and ingredients are required!', 'danger')
            return redirect(url_for('add_meal'))
        
        try:
            calories = int(calories) if calories else 0
            protein = int(protein) if protein else 0
            carbs = int(carbs) if carbs else 0
            fat = int(fat) if fat else 0
        except ValueError:
            flash('Nutrition values must be numbers!', 'danger')
            return redirect(url_for('add_meal'))
        
        meal = {
            'id': get_next_id(),
            'name': name,
            'category': category,
            'ingredients': ingredients,
            'calories': calories,
            'protein': protein,
            'carbs': carbs,
            'fat': fat,
            'instructions': instructions,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        save_meal(meal)
        flash('Meal added successfully!', 'success')
        return redirect(url_for('index'))
    
    return render_template('add_meal.html')

@app.route('/meal/<int:meal_id>')
def view_meal(meal_id):
    meal = get_meal_by_id(meal_id)
    if not meal:
        flash('Meal not found!', 'danger')
        return redirect(url_for('index'))
    return render_template('view_meal.html', meal=meal)

@app.route('/edit/<int:meal_id>', methods=['GET', 'POST'])
def edit_meal(meal_id):
    meal = get_meal_by_id(meal_id)
    if not meal:
        flash('Meal not found!', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        name = request.form['name']
        category = request.form['category']
        ingredients = request.form['ingredients']
        calories = request.form['calories']
        protein = request.form['protein']
        carbs = request.form['carbs']
        fat = request.form['fat']
        instructions = request.form['instructions']
        
        if not name or not category or not ingredients:
            flash('Name, category, and ingredients are required!', 'danger')
            return redirect(url_for('edit_meal', meal_id=meal_id))
        
        try:
            calories = int(calories) if calories else 0
            protein = int(protein) if protein else 0
            carbs = int(carbs) if carbs else 0
            fat = int(fat) if fat else 0
        except ValueError:
            flash('Nutrition values must be numbers!', 'danger')
            return redirect(url_for('edit_meal', meal_id=meal_id))
        
        updated_meal = {
            'id': meal_id,
            'name': name,
            'category': category,
            'ingredients': ingredients,
            'calories': calories,
            'protein': protein,
            'carbs': carbs,
            'fat': fat,
            'instructions': instructions,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        update_meal(meal_id, updated_meal)
        flash('Meal updated successfully!', 'success')
        return redirect(url_for('view_meal', meal_id=meal_id))
    
    return render_template('edit_meal.html', meal=meal)

@app.route('/delete/<int:meal_id>', methods=['POST'])
def delete_meal_route(meal_id):
    meal = get_meal_by_id(meal_id)
    if not meal:
        flash('Meal not found!', 'danger')
        return redirect(url_for('index'))
    
    delete_meal(meal_id)
    flash('Meal deleted successfully!', 'success')
    return redirect(url_for('index'))

@app.route('/plan', methods=['GET', 'POST'])
def meal_plan():
    meals = get_all_meals()
    selected_meals = []
    totals = None
    
    if request.method == 'POST':
        selected_meal_ids = request.form.getlist('meal_ids')
        selected_meals = [meal for meal in meals if meal['id'] in selected_meal_ids]
        totals = calculate_nutrition_totals(selected_meal_ids)
    
    return render_template('meal_plan.html', meals=meals, selected_meals=selected_meals, totals=totals)

if _name_ == '_main_':
    init_excel_file()
    app.run(debug=True)
